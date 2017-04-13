from openpyxl import load_workbook
import pickle
from pprint import pprint

wb = load_workbook("scans.xlsx", read_only=True)
ws = wb['Sheet1']

## read headers from row 1
headers = {}
for cell in ws[1]:
    headers[str(cell.value)] = int(cell.column)-1
print "headers:" 
pprint(headers)

filenames = {}
samples = {}
orimats = {}

for row in ws.iter_rows(min_row=2):
    run = int(row[headers['run #']].value)
    filename = str(row[headers['filename']].value)
    sample = str(row[headers['sample']].value)
    orimat = str(row[headers['orientation matrix']].value)

    filenames[run] = filename
    samples[run] = sample 
    orimats[run] = orimat 

pickle.dump({"filenames":filenames, 'samples':samples, \
             "orimats":orimats}, \
            open("scans.pickle", 'wb'))
